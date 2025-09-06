# File: /opt/access_control/src/api/web_api.py
# Dashboard web modulare - API principale - FIXED IMPORTS

import os
import sys
import json
import sqlite3
import logging
import shutil
import subprocess
import psutil
import threading
from datetime import datetime
from typing import Dict, Any, Tuple, Optional, List, Union, Callable

from flask import Flask, render_template_string, request, jsonify, session, redirect, send_from_directory, Response, Blueprint, current_app, make_response
from flask_cors import CORS
from werkzeug.exceptions import BadRequest, InternalServerError
from werkzeug.wrappers import Response as WerkzeugResponse
from werkzeug.datastructures import ImmutableMultiDict
from werkzeug.local import LocalProxy

# Assicurati che il path sia corretto
sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))

# Configura il fuso orario per l'applicazione
import pytz
import time
os.environ['TZ'] = 'Europe/Rome'
try:
    time.tzset()  # Disponibile solo su Unix
except AttributeError:
    pass  # Su Windows non è disponibile

# Importazioni locali
from api.auth import verify_user, USER_ROLES
from api.utils import get_db_connection, require_auth, require_permission
import api.hardware_tests as hardware_tests
import api.backup_module as backup_module
import api.hardware_detection as hardware_detection
from core.config import get_config_manager
from api.dashboard_templates import LOGIN_TEMPLATE, get_dashboard_template, ADMIN_CONFIG_TEMPLATE, ADMIN_BACKUP_TEMPLATE

# Importazioni hardware
# CardReader importato dinamicamente per supportare sia CRT-285 che Omnikey
from hardware.usb_rly08_controller import USBRLY08Controller
from external.odoo_partner_connector import OdooPartnerConnector

# Importazioni dei moduli
from api.modules.profilo import profilo_bp
from api.modules.user_management import user_management_bp
from api.modules.log_management import log_management_bp
from api.modules.utenti_autorizzati import utenti_autorizzati_bp
from api.modules.system_users import system_users_bp
from api.modules.activities import activities_bp
from api.modules.configurazione_accessi import configurazione_accessi_bp, verifica_orario, verifica_limite_mensile
from api.backup_module import backup_bp

# Definizione dei tipi personalizzati
FlaskResponse = Union[Response, str, Tuple[Union[Dict[str, Any], str], int]]
Request = Union[ImmutableMultiDict[str, str], Dict[str, Any]]

# Configurazione del logging
logging.basicConfig(level=logging.DEBUG, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# Queste importazioni sono già state definite sopra e non sono necessarie qui

# Questi import sono già stati definiti sopra e non sono necessari qui

from typing import Optional

# Definizione dei tipi personalizzati
FlaskResponse = Union[Response, str, Tuple[Union[Dict[str, Any], str], int]]
Request = Union[ImmutableMultiDict[str, str], Dict[str, Any]]

# Gestione delle eccezioni
def handle_exception(e: Exception) -> Tuple[Dict[str, Any], int]:
    logger.error(f"Errore: {str(e)}", exc_info=True)
    if isinstance(e, BadRequest):
        return jsonify({"success": False, "error": "Richiesta non valida"}), 400
    elif isinstance(e, InternalServerError):
        return jsonify({"success": False, "error": "Errore interno del server"}), 500
    else:
        return jsonify({"success": False, "error": str(e)}), 500

# Definizione dei tipi per le funzioni di Flask
def flask_route(rule: str, **options: Any) -> Callable[[Callable[..., FlaskResponse]], Callable[..., FlaskResponse]]:
    return app.route(rule, **options)

def flask_jsonify(*args: Any, **kwargs: Any) -> WerkzeugResponse:
    return jsonify(*args, **kwargs)

# Tipo per le funzioni di route
RouteFunction = Callable[..., FlaskResponse]

app = Flask(__name__)
app.secret_key = 'raee-2025-access-control-system'

# Configura CORS per permettere richieste cross-origin con credenziali
CORS(app, 
     supports_credentials=True,
     resources={r"/*": {"origins": "*"}},
     allow_headers=["Content-Type", "Authorization"],
     expose_headers=["Content-Type", "X-CSRFToken"],
     methods=["GET", "POST", "PUT", "DELETE", "OPTIONS"])

# Configura sessione per funzionare cross-origin
app.config.update(
    SESSION_COOKIE_SECURE=False,  # Impostare a True se si usa HTTPS
    SESSION_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SAMESITE='Lax'
)

project_root = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', '..'))
sys.path.insert(0, project_root)

# Database path persistente
DB_PATH = '/opt/access_control/src/access.db'

# IMPORTA I MODULI DOPO aver definito le funzioni condivise
from api.modules.profilo import profilo_bp
from api.modules.user_management import user_management_bp
from api.modules.log_management import log_management_bp
from api.modules.utenti_autorizzati import utenti_autorizzati_bp
from api.modules.system_users import system_users_bp
from api.modules.activities import activities_bp
from api.modules.configurazione_accessi import configurazione_accessi_bp

# Route di test per visualizzare tutte le routes
@app.route('/api/routes')
def list_routes():
    """Lista tutte le routes registrate"""
    routes = []
    for rule in app.url_map.iter_rules():
        routes.append({
            'endpoint': rule.endpoint,
            'methods': sorted(list(rule.methods)),
            'path': str(rule)
        })
    return jsonify({
        'success': True,
        'routes': sorted(routes, key=lambda x: x['path'])
    })

# Registra i blueprint
app.register_blueprint(profilo_bp)
app.register_blueprint(user_management_bp)
app.register_blueprint(log_management_bp)
app.register_blueprint(utenti_autorizzati_bp)
app.register_blueprint(system_users_bp)
app.register_blueprint(activities_bp)
app.register_blueprint(configurazione_accessi_bp)
app.register_blueprint(backup_bp)

# ===============================
# ROUTES PRINCIPALI
# ===============================

@flask_route('/')
@require_auth()
def dashboard() -> FlaskResponse:
    """Dashboard principale con template personalizzato per ruolo"""
    # Menu items per admin
    menu_items: List[Dict[str, str]] = []
    if session.get('role') == 'admin':
        menu_items = [
            {
                'url': '/utenti-autorizzati',
                'icon': 'fas fa-users',
                'text': 'Utenti Autorizzati'
            },
            {
                'url': '/admin/users',
                'icon': 'fas fa-users-cog',
                'text': 'Gestione Utenti Sistema'
            },
            {
                'url': '/configurazione-orari',
                'icon': 'fas fa-clock',
                'text': 'Configurazione Orari'
            },
            {
                'url': '/test-accessi',
                'icon': 'fas fa-door-open',
                'text': 'Test Accessi'
            }
        ]
    
    return render_template_string(
        get_dashboard_template(), 
        session=session,
        menu_items=menu_items
    )

@app.route('/utenti-autorizzati')
@require_auth()
@require_permission('all')
def utenti_autorizzati_page() -> FlaskResponse:
    """Pagina gestione utenti autorizzati"""
    with open(os.path.join(os.path.dirname(__file__), 'templates', 'utenti_autorizzati.html'), 'r') as f:
        template = f.read()
    return render_template_string(template, session=session)

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '').strip()
        
        logger.debug(f"Tentativo di login per l'utente: {username}")
        
        if not username or not password:
            logger.debug("Username o password mancanti")
            return render_template_string(LOGIN_TEMPLATE, error="Inserire username e password")
        
        # Verifica credenziali nel database
        success, role = verify_user(username, password)
        
        logger.debug(f"Risultato verifica utente: success={success}, role={role}")
        
        if success:
            conn = get_db_connection()
            if conn:
                try:
                    cursor = conn.cursor()
                    # Aggiorna last_login
                    cursor.execute("""
                        UPDATE utenti_sistema 
                        SET last_login = CURRENT_TIMESTAMP,
                            modified_at = CURRENT_TIMESTAMP,
                            modified_by = 'system'
                        WHERE username = ?
                    """, (username,))
                    
                    # Log accesso
                    cursor.execute("""
                        INSERT INTO eventi_sistema (tipo_evento, livello, messaggio, componente)
                        VALUES (?, ?, ?, ?)
                    """, ('LOGIN', 'INFO', f'Login utente {username} ({role})', 'AUTH'))
                    
                    conn.commit()
                    logger.debug("Aggiornamento last_login e log accesso completati")
                except Exception as e:
                    logger.error(f"Errore durante l'aggiornamento del database: {e}")
                    conn.rollback()
                finally:
                    conn.close()

            # Imposta dati sessione
            session['username'] = username
            session['role'] = role
            session['logged_in'] = True
            session['login_time'] = datetime.now().strftime('%d/%m/%Y %H:%M')
            session['role_name'] = USER_ROLES.get(role, {}).get('name', 'Utente')
            session['permissions'] = USER_ROLES.get(role, {}).get('permissions', [])
            
            logger.debug("Dati sessione impostati correttamente")
            logger.debug(f"Sessione attuale: {session}")
            
            return redirect('/')
        else:
            logger.debug("Autenticazione fallita")
            return render_template_string(LOGIN_TEMPLATE, error="Credenziali non valide")
    
    # Mostra info sui livelli nel form di login
    login_info = """
    <div class="mt-4 text-center">
        <small class="text-muted">
            <strong>Utenti di test:</strong><br>
            🔴 admin/admin123 (Amministratore)<br>
            🔵 manager/manager123 (Gestore Utenti)<br>
            🟢 viewer/viewer123 (Visualizzatore)
        </small>
    </div>
    """
    
    return render_template_string(LOGIN_TEMPLATE + login_info)

@app.route('/logout')
def logout():
    """Gestisce il logout utente con gestione errori robusta"""
    try:
        username = session.get('username')
        role = session.get('role')
        
        # Log logout se possibile
        if username:
            try:
                conn = get_db_connection()
                if conn:
                    try:
                        cursor = conn.cursor()
                        cursor.execute("""
                            INSERT INTO eventi_sistema (tipo_evento, livello, messaggio, componente)
                            VALUES (?, ?, ?, ?)
                        """, ('LOGOUT', 'INFO', f'Logout utente {username} ({role})', 'AUTH'))
                        conn.commit()
                    except Exception as e:
                        logger.error(f"Errore logging logout: {e}")
                    finally:
                        conn.close()
            except Exception as e:
                logger.error(f"Errore connessione DB durante logout: {e}")
        
        # Pulisci la sessione in ogni caso
        session.clear()
        
        return redirect('/login')
        
    except Exception as e:
        logger.error(f"Errore durante logout: {e}")
        # Forza pulizia sessione anche in caso di errore
        session.clear()
        return redirect('/login')

@app.route('/api/profile/info')
@require_auth()
def api_profile_info():
    """Restituisce informazioni profilo dalla sessione"""
    return jsonify({
        'success': True,
        'user': {
            'username': session.get('username'),
            'role': session.get('role'),
            'login_time': session.get('login_time', datetime.now().strftime('%d/%m/%Y %H:%M'))
        }
    })

@app.route('/api/user-menu-html')
@require_auth()
def api_user_menu_html():
    """Restituisce HTML del menu utente con dati sessione"""
    
    # Mappa ruoli aggiornata per 3 livelli
    role_names = {
        'admin': 'Amministratore',
        'user_manager': 'Gestore Utenti', 
        'viewer': 'Visualizzatore'
    }
    
    user_role = session.get('role', 'viewer')
    user_role_name = role_names.get(user_role, 'Utente')
    username = session.get('username', 'user')
    
    # Template del menu inline
    menu_html = f'''
    <div class="user-menu-container">
        <!-- Trigger Menu -->
        <div class="user-menu-trigger">
            <div class="user-avatar">
                {username[0].upper()}
            </div>
            <div class="user-menu-info">
                <div class="user-menu-name">{username}</div>
                <div class="user-menu-role">{user_role_name}</div>
            </div>
            <i class="fas fa-chevron-down" style="color: white; font-size: 0.8rem;"></i>
        </div>
        
        <!-- Dropdown Menu -->
        <div class="user-menu-dropdown">
            <div class="user-menu-header">
                <div class="avatar-large">
                    {username[0].upper()}
                </div>
                <div class="user-name">{username}</div>
                <div class="user-email">{user_role_name}</div>
            </div>
            
            <div class="user-menu-items">
                <a href="#" class="user-menu-item" data-action="profile">
                    <i class="fas fa-user-circle"></i>
                    <span class="user-menu-item-text">Il Mio Profilo</span>
                </a>
                
                <a href="#" class="user-menu-item" data-action="change-password">
                    <i class="fas fa-key"></i>
                    <span class="user-menu-item-text">Cambio Password</span>
                </a>
                
                <div class="user-menu-divider"></div>
                
                <a href="/logout" class="user-menu-item logout">
                    <i class="fas fa-sign-out-alt"></i>
                    <span class="user-menu-item-text">Logout</span>
                </a>
            </div>
        </div>
    </div>
    
    <!-- Overlay -->
    <div class="user-menu-overlay"></div>
    '''
    
    return menu_html

# ===============================
# STATIC FILES
# ===============================

@app.route('/static/<path:path>')
def send_static(path):
    return send_from_directory(os.path.join(os.path.dirname(__file__), 'static'), path)

# ===============================
# API ENDPOINTS - AUTORIZZAZIONE
# ===============================

@app.route('/api/authorize', methods=['POST'])
def authorize():
    """Gestisce l'autorizzazione dell'accesso quando viene passato il codice fiscale"""
    try:
        data = request.get_json()
        if not data or 'codice_fiscale' not in data:
            return jsonify({'success': False, 'error': 'Codice fiscale mancante'}), 400

        codice_fiscale = data['codice_fiscale']
        logger.info(f"Richiesta autorizzazione per CF: {codice_fiscale}")

        # Processa il codice fiscale con la stessa logica usata per la lettura da Omnikey
        result = process_codice_fiscale(codice_fiscale)
        
        if result['authorized']:
            return jsonify({
                'success': True,
                'message': 'Accesso autorizzato',
                'user': result['user_name']
            })
        else:
            return jsonify({
                'success': False,
                'error': result['error_message']
            }), 403

    except Exception as e:
        logger.error(f"Errore generico autorizzazione: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500

def get_nome_installazione():
    """Recupera il nome installazione dal database"""
    try:
        conn = get_db_connection()
        if conn:
            cursor = conn.cursor()
            cursor.execute("SELECT value FROM system_settings WHERE key = 'sistema.nome_installazione'")
            result = cursor.fetchone()
            conn.close()
            if result:
                return result[0]
    except Exception as e:
        logger.error(f"Errore recupero nome installazione: {e}")
    return "Terminale"  # Default fallback

def process_codice_fiscale(codice_fiscale):
    """
    Processa un codice fiscale per autorizzazione accesso.
    Questa funzione è usata sia dall'endpoint API che dal lettore Omnikey.
    """
    start_time = time.time()
    result = {
        'authorized': False,
        'user_name': None,
        'error_message': None,
        'tipo_accesso': 'NEGATO'  # AUTORIZZATO, NEGATO, FUORI_ORARIO, LIMITE_SUPERATO, UTENTE_DISATTIVATO
    }
    
    # Maschera il CF per i log
    masked_cf = f"{codice_fiscale[:4]}***{codice_fiscale[-4:]}"
    logger.info(f"Elaborazione CF: {masked_cf}")
    
    # Variabili per il logging
    processing_time = time.time() - start_time
    tipo_accesso = 'NEGATO'
    nome_utente = None
    
    # Recupera il nome dell'installazione
    nome_terminale = get_nome_installazione()
    
    try:
        # Connessione al database per registrare TUTTI gli accessi
        conn = get_db_connection()
        if not conn:
            result['error_message'] = 'Database non disponibile'
            return result
        
        cursor = conn.cursor()
        
        # 1. Prima verifica se l'utente esiste nel database
        try:
            cursor.execute('''
                SELECT id, codice_fiscale, nome, attivo
                FROM utenti_autorizzati 
                WHERE codice_fiscale = ?
            ''', (codice_fiscale,))
            
            user = cursor.fetchone()
            if not user:
                error_msg = 'Utente non trovato'
                logger.warning(f"Accesso negato per CF {masked_cf}: utente non trovato")
                result['error_message'] = error_msg
                result['tipo_accesso'] = 'UTENTE_NON_TROVATO'
                
                # Registra accesso NEGATO per utente non trovato
                cursor.execute('''
                    INSERT INTO log_accessi 
                    (timestamp, codice_fiscale, autorizzato, durata_elaborazione, terminale_id, motivo_rifiuto, tipo_accesso)
                    VALUES (CURRENT_TIMESTAMP, ?, 0, ?, ?, ?, ?)
                ''', (codice_fiscale, processing_time, nome_terminale, error_msg, 'UTENTE_NON_TROVATO'))
                conn.commit()
                conn.close()
                return result
            
            nome_utente = user[2]
            
            if not user[3]:  # not attivo
                error_msg = 'Utente disattivato'
                logger.warning(f"Accesso negato per CF {masked_cf}: utente disattivato")
                result['error_message'] = error_msg
                result['tipo_accesso'] = 'UTENTE_DISATTIVATO'
                
                # Registra accesso NEGATO per utente disattivato
                cursor.execute('''
                    INSERT INTO log_accessi 
                    (timestamp, codice_fiscale, autorizzato, durata_elaborazione, terminale_id, motivo_rifiuto, nome_utente, tipo_accesso)
                    VALUES (CURRENT_TIMESTAMP, ?, 0, ?, ?, ?, ?, ?)
                ''', (codice_fiscale, processing_time, nome_terminale, error_msg, nome_utente, 'UTENTE_DISATTIVATO'))
                conn.commit()
                conn.close()
                return result

            # 2. Solo se l'utente esiste ed è attivo, verifica l'orario
            if not verifica_orario():
                error_msg = 'Accesso non consentito in questo orario'
                logger.warning(f"Accesso negato per CF {masked_cf}: fuori orario")
                result['error_message'] = error_msg
                result['tipo_accesso'] = 'FUORI_ORARIO'
                
                # Registra accesso NEGATO per fuori orario CON il nome utente
                cursor.execute('''
                    INSERT INTO log_accessi 
                    (timestamp, codice_fiscale, autorizzato, durata_elaborazione, terminale_id, motivo_rifiuto, nome_utente, tipo_accesso)
                    VALUES (CURRENT_TIMESTAMP, ?, 0, ?, ?, ?, ?, ?)
                ''', (codice_fiscale, processing_time, nome_terminale, error_msg, nome_utente, 'FUORI_ORARIO'))
                conn.commit()
                conn.close()
                return result

            # 3. Verifica limite mensile
            if not verifica_limite_mensile(codice_fiscale):
                error_msg = 'Limite mensile accessi superato'
                logger.warning(f"Accesso negato per CF {masked_cf}: limite mensile superato")
                result['error_message'] = error_msg
                result['tipo_accesso'] = 'LIMITE_SUPERATO'
                
                # Registra accesso NEGATO per limite superato
                cursor.execute('''
                    INSERT INTO log_accessi 
                    (timestamp, codice_fiscale, autorizzato, durata_elaborazione, terminale_id, motivo_rifiuto, nome_utente, tipo_accesso)
                    VALUES (CURRENT_TIMESTAMP, ?, 0, ?, ?, ?, ?, ?)
                ''', (codice_fiscale, processing_time, nome_terminale, error_msg, nome_utente, 'LIMITE_SUPERATO'))
                conn.commit()
                conn.close()
                return result

            # 4. ACCESSO AUTORIZZATO - Incrementa contatore accessi
            oggi = datetime.now()
            cursor.execute('''
                INSERT OR REPLACE INTO conteggio_ingressi_mensili 
                (codice_fiscale, mese, anno, numero_ingressi, ultimo_ingresso)
                VALUES (
                    ?,
                    ?,
                    ?,
                    COALESCE(
                        (SELECT numero_ingressi + 1 
                         FROM conteggio_ingressi_mensili 
                         WHERE codice_fiscale = ? AND mese = ? AND anno = ?),
                        1
                    ),
                    CURRENT_TIMESTAMP
                )
            ''', (codice_fiscale, oggi.month, oggi.year, codice_fiscale, oggi.month, oggi.year))

            # 5. Log accesso AUTORIZZATO
            cursor.execute('''
                INSERT INTO log_accessi 
                (timestamp, codice_fiscale, autorizzato, durata_elaborazione, terminale_id, nome_utente, tipo_accesso)
                VALUES (CURRENT_TIMESTAMP, ?, 1, ?, ?, ?, ?)
            ''', (codice_fiscale, processing_time, nome_terminale, nome_utente, 'AUTORIZZATO'))

            conn.commit()
            conn.close()

            # 6. Imposta risultato positivo
            result['authorized'] = True
            result['user_name'] = nome_utente
            result['tipo_accesso'] = 'AUTORIZZATO'
            logger.info(f"Accesso autorizzato per {result['user_name']} (CF: {masked_cf})")
            
            return result

        except Exception as e:
            logger.error(f"Errore durante autorizzazione per CF {masked_cf}: {str(e)}")
            result['error_message'] = str(e)
            
            # Registra anche gli errori
            try:
                cursor.execute('''
                    INSERT INTO log_accessi 
                    (timestamp, codice_fiscale, autorizzato, durata_elaborazione, terminale_id, motivo_rifiuto, tipo_accesso)
                    VALUES (CURRENT_TIMESTAMP, ?, 0, ?, ?, ?, ?)
                ''', (codice_fiscale, processing_time, nome_terminale, str(e), 'ERRORE'))
                conn.commit()
            except:
                pass
            
            if conn:
                conn.close()
            return result
        finally:
            conn.close()

    except Exception as e:
        logger.error(f"Errore generico autorizzazione: {str(e)}")
        result['error_message'] = str(e)
        return result

# ===============================
# API ENDPOINTS - DATI
# ===============================

@app.route('/api/stats')
@require_auth()
def api_stats():
    """Statistiche usando schema database_manager.py"""
    conn = get_db_connection()
    if not conn:
        return jsonify({'error': 'Database non disponibile'}), 500
    
    try:
        cursor = conn.cursor()
        
        # Oggi
        today = datetime.now().strftime('%Y-%m-%d')
        
        # Accessi oggi
        cursor.execute("""
            SELECT COUNT(*) 
            FROM log_accessi 
            WHERE DATE(timestamp) = ?
        """, (today,))
        accessi_oggi = cursor.fetchone()[0]
        
        # Accessi autorizzati oggi
        cursor.execute("""
            SELECT COUNT(*) 
            FROM log_accessi 
            WHERE DATE(timestamp) = ? AND autorizzato = 1
        """, (today,))
        autorizzati_oggi = cursor.fetchone()[0]
        
        return jsonify({
            'accessi_oggi': accessi_oggi,
            'autorizzati_oggi': autorizzati_oggi
        })
        
    except Exception as e:
        print(f"Errore API stats: {e}")
        return jsonify({'error': str(e)}), 500
    finally:
        conn.close()

@app.route('/api/users/count')
@require_auth()
def api_users_count():
    """Count utenti usando schema database_manager.py"""
    conn = get_db_connection()
    if not conn:
        return jsonify({'error': 'Database non disponibile'}), 500
    
    try:
        cursor = conn.cursor()
        cursor.execute("SELECT COUNT(*) FROM utenti_autorizzati WHERE attivo = 1")
        count = cursor.fetchone()[0]
        return jsonify({'count': count})
    except Exception as e:
        print(f"Errore API users count: {e}")
        return jsonify({'error': str(e)}), 500
    finally:
        conn.close()

@app.route('/api/recent-accesses')
@require_auth()
def api_recent_accesses():
    """Accessi recenti usando schema database_manager.py"""
    limit = request.args.get('limit', 10, type=int)
    
    conn = get_db_connection()
    if not conn:
        return jsonify({'error': 'Database non disponibile'}), 500
    
    try:
        cursor = conn.cursor()
        cursor.execute("""
            SELECT 
                la.timestamp, 
                la.codice_fiscale, 
                la.autorizzato,
                COALESCE(ua.nome, 'Utente sconosciuto') as nome_completo
            FROM log_accessi la
            LEFT JOIN utenti_autorizzati ua ON la.codice_fiscale = ua.codice_fiscale
            ORDER BY la.timestamp DESC
            LIMIT ?
        """, (limit,))
        
        results = cursor.fetchall()
        accesses = []
        
        for row in results:
            accesses.append({
                'timestamp': row[0],
                'codice_fiscale': row[1],
                'autorizzato': bool(row[2]),
                'nome': row[3]
            })
        
        return jsonify({'accesses': accesses})
        
    except Exception as e:
        print(f"Errore API recent accesses: {e}")
        return jsonify({'error': str(e)}), 500
    finally:
        conn.close()

# ===============================
# API ENDPOINTS - TEST HARDWARE
# ===============================

@app.route('/api/test/gate')
@require_auth()
def api_test_gate():
    """Test cancello"""
    return hardware_tests.test_gate(get_db_connection)

@app.route('/api/test_relay', methods=['POST'])
@require_auth()
def api_test_relay():
    """Test USB-RLY08"""
    return hardware_tests.test_relay()

@app.route('/api/test/integrated', methods=['POST'])
@require_auth()
def api_test_integrated():
    """Test integrato completo"""
    return hardware_tests.test_integrated(get_db_connection)

@app.route('/api/hardware/test-reader', methods=['POST'])
@require_auth()
def api_hardware_test_reader():
    """Test lettore tessere - CORRETTA CHIAMATA A hardware_tests"""
    return hardware_tests.test_reader()

@app.route('/api/hardware/stop-reader', methods=['POST'])
@require_auth()
def api_hardware_stop_reader():
    """Ferma il test/monitor del lettore"""
    return hardware_tests.stop_reader()

# ===============================
# API ENDPOINTS - ODOO SYNC
# ===============================

@app.route('/api/odoo/sync', methods=['POST'])
@require_auth()
@require_permission('all')
def api_odoo_sync():
    """Forza la sincronizzazione con Odoo"""
    try:
        success, stats = perform_odoo_sync()
        
        if success:
            return jsonify({
                'success': True,
                'message': f"Sincronizzazione completata: {stats['added']} cittadini aggiunti",
                'stats': stats
            })
        else:
            return jsonify({
                'success': False,
                'error': "Sincronizzazione fallita",
                'stats': stats
            }), 500
            
    except Exception as e:
        logger.error(f"Errore API sync Odoo: {e}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/api/odoo/status')
@require_auth()
def api_odoo_status():
    """Stato connessione Odoo"""
    try:
        if not odoo_connector:
            return jsonify({
                'success': False,
                'error': "Connettore Odoo non configurato"
            }), 404
        
        status = odoo_connector.get_sync_status()
        
        return jsonify({
            'success': True,
            'status': status
        })
        
    except Exception as e:
        logger.error(f"Errore API status Odoo: {e}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

# ===============================
# API ENDPOINTS - STATI
# ===============================

@app.route('/api/relay_status')
@require_auth()
def api_relay_status():
    """Stato test relay"""
    return hardware_tests.get_relay_status()

@app.route('/api/integrated_status')
@require_auth()
def api_integrated_status():
    """Stato test integrato"""
    return hardware_tests.get_integrated_status()

@app.route('/api/hardware/status')
@require_auth()
def api_hardware_status():
    """Stato test hardware"""
    test_id = request.args.get('test_id')
    return hardware_tests.get_hardware_status(test_id)

# ===============================
# API ENDPOINTS - RILEVAMENTO HARDWARE
# ===============================

@app.route('/api/hardware/detect')
@require_auth()
@require_permission('all')
def api_hardware_detect():
    """Rileva automaticamente l'hardware collegato"""
    return jsonify(hardware_detection.detect_hardware())

@app.route('/api/hardware/test-connection', methods=['POST'])
@require_auth()
@require_permission('all')
def api_hardware_test_connection():
    """Testa la connessione a un dispositivo hardware"""
    if not request.is_json:
        return jsonify({'success': False, 'error': 'Richiesta non valida, JSON richiesto'})
    
    data = request.get_json()
    hardware_type = data.get('type')
    device_path = data.get('device_path', data.get('port'))  # Supporta sia 'device_path' che 'port' per retrocompatibilità
    
    if not hardware_type or not device_path:
        return jsonify({
            'success': False,
            'error': 'Tipo di hardware e porta sono richiesti'
        })
    
    # Estrai parametri aggiuntivi specifici per tipo di hardware
    kwargs = {}
    
    # Per lettore tessere
    if hardware_type == 'card_reader':
        if 'reader_type' in data:
            kwargs['reader_type'] = data['reader_type']
    
    # Per controller relè
    elif hardware_type == 'relay':
        if 'baud_rate' in data:
            kwargs['baud_rate'] = data['baud_rate']
    
    # Log parametri
    logger.info(f"Test connessione hardware: tipo={hardware_type}, device_path={device_path}, kwargs={kwargs}")
    
    # Esegui test connessione con parametri aggiuntivi
    return jsonify(hardware_detection.test_hardware_connection(hardware_type, device_path, **kwargs))

@app.route('/api/hardware/config')
@require_auth()
@require_permission('all')
def api_hardware_config():
    """Carica la configurazione dell'hardware"""
    return jsonify(hardware_detection.load_hardware_config(get_db_connection))

@app.route('/api/hardware/config/save', methods=['POST'])
@require_auth()
@require_permission('all')
def api_hardware_config_save():
    """Salva la configurazione dell'hardware"""
    data = request.get_json()
    return jsonify(hardware_detection.save_hardware_config(data, get_db_connection))

# ===============================
# ROUTES GESTIONE SISTEMA ADMIN
# ===============================

@app.route('/admin/users')
@require_auth()
@require_permission('all')  # Solo admin
def admin_users():
    """Gestione completa utenti - Solo Admin"""
    with open(os.path.join(os.path.dirname(__file__), 'static', 'html', 'user-management.html'), 'r') as f:
        template = f.read()
    return render_template_string(template, session=session)

@app.route('/admin/config')
@require_auth()
@require_permission('all')  # Solo admin
def admin_config():
    """Configurazioni sistema - Solo Admin"""
    return render_template_string(ADMIN_CONFIG_TEMPLATE, session=session)

@app.route('/log-accessi')
@require_auth()
def log_accessi_page():
    """Pagina Log Accessi"""
    from log_accessi_template import LOG_ACCESSI_TEMPLATE
    return LOG_ACCESSI_TEMPLATE

@app.route('/admin/backup')
@require_auth()
@require_permission('all')  # Solo admin
def admin_backup():
    """Backup & Restore - Solo Admin"""
    return render_template_string(ADMIN_BACKUP_TEMPLATE, session=session)

@app.route('/api/admin/clock-config', methods=['GET'])
@require_auth()
@require_permission('all')  # Solo admin
def api_get_clock_config():
    """Recupera configurazioni orologio"""
    try:
        conn = get_db_connection()
        if conn:
            cursor = conn.cursor()
            # Recupera configurazioni orologio
            config = {}
            keys = ['sistema.timezone', 'sistema.formato_data', 'sistema.formato_ora', 
                    'sistema.ntp_enabled', 'sistema.ntp_server']
            
            for key in keys:
                cursor.execute("SELECT value FROM system_settings WHERE key = ?", (key,))
                result = cursor.fetchone()
                simple_key = key.replace('sistema.', '')
                if result:
                    value = result[0]
                    # Converti booleani
                    if simple_key == 'ntp_enabled':
                        config[simple_key] = value.lower() == 'true'
                    else:
                        config[simple_key] = value
                else:
                    # Valori default
                    if simple_key == 'timezone':
                        config[simple_key] = 'Europe/Rome'
                    elif simple_key == 'formato_data':
                        config[simple_key] = 'DD/MM/YYYY'
                    elif simple_key == 'formato_ora':
                        config[simple_key] = '24'
                    elif simple_key == 'ntp_enabled':
                        config[simple_key] = True
                    elif simple_key == 'ntp_server':
                        config[simple_key] = 'pool.ntp.org'
            
            conn.close()
            return jsonify({'success': True, 'config': config})
    except Exception as e:
        logger.error(f"Errore recupero config orologio: {e}")
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/admin/clock-config', methods=['POST'])
@require_auth()
@require_permission('all')  # Solo admin
def api_save_clock_config():
    """Salva configurazioni orologio"""
    try:
        data = request.get_json()
        conn = get_db_connection()
        if conn:
            cursor = conn.cursor()
            
            # Mappa dei campi da salvare
            settings = {
                'sistema.timezone': data.get('timezone', 'Europe/Rome'),
                'sistema.formato_data': data.get('formato_data', 'DD/MM/YYYY'),
                'sistema.formato_ora': data.get('formato_ora', '24'),
                'sistema.ntp_enabled': 'true' if data.get('ntp_enabled') else 'false',
                'sistema.ntp_server': data.get('ntp_server', 'pool.ntp.org')
            }
            
            # Salva o aggiorna ogni impostazione
            for key, value in settings.items():
                cursor.execute("""
                    INSERT OR REPLACE INTO system_settings (key, value, updated_at) 
                    VALUES (?, ?, datetime('now'))
                """, (key, value))
            
            conn.commit()
            conn.close()
            
            logger.info(f"Configurazioni orologio salvate: {settings}")
            return jsonify({'success': True})
            
    except Exception as e:
        logger.error(f"Errore salvataggio config orologio: {e}")
        return jsonify({'success': False, 'error': str(e)})

@app.route('/configurazione-orari')
@require_auth()
@require_permission('all')
def configurazione_orari():
    """Pagina configurazione orari accesso"""
    with open(os.path.join(os.path.dirname(__file__), 'templates', 'configurazione_orari.html'), 'r') as f:
        template = f.read()
    return render_template_string(template, session=session)

@app.route('/test-accessi')
@require_auth()
@require_permission('all')
def test_accessi():
    """Pagina test funzionalità accessi"""
    with open(os.path.join(os.path.dirname(__file__), 'static', 'html', 'test-accessi.html'), 'r') as f:
        template = f.read()
    return render_template_string(template, session=session)
# ===============================
# API ENDPOINTS BACKUP (utilizzando backup_module esistente)
# ===============================

@app.route('/api/backup/status')
@require_auth()
@require_permission('all')
def api_backup_status():
    return backup_module.get_backup_status()

@app.route('/api/backup/create', methods=['POST'])
@require_auth()
@require_permission('all')
def api_backup_create():
    data = request.get_json()
    backup_type = data.get('type', 'complete')
    return backup_module.create_backup(backup_type)

@app.route('/api/backup/cleanup', methods=['POST'])
@require_auth()
@require_permission('all')
def api_backup_cleanup():
    return backup_module.cleanup_old_backups()

@app.route('/api/backup/schedule', methods=['POST'])
@require_auth()
@require_permission('all')
def api_backup_schedule():
    """Salva configurazione schedulazione backup"""
    try:
        data = request.get_json()
        scheduling = data.get('scheduling', 'daily')
        retention = int(data.get('retention', 7))
        conn = get_db_connection()
        if not conn:
            return jsonify({'success': False, 'error': 'Database non disponibile'}), 500
        try:
            cursor = conn.cursor()
            cursor.execute('''
                INSERT OR REPLACE INTO system_settings (key, value, updated_at)
                VALUES (?, ?, CURRENT_TIMESTAMP)
            ''', ('backup.scheduling', scheduling))
            cursor.execute('''
                INSERT OR REPLACE INTO system_settings (key, value, updated_at)
                VALUES (?, ?, CURRENT_TIMESTAMP)
            ''', ('backup.retention', str(retention)))
            conn.commit()
            return jsonify({'success': True, 'message': 'Schedulazione backup salvata'})
        finally:
            conn.close()
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/backup/schedule')
@require_auth()
@require_permission('all')
def api_get_backup_schedule():
    """Restituisce configurazione schedulazione backup"""
    try:
        conn = get_db_connection()
        if not conn:
            return jsonify({'success': False, 'error': 'Database non disponibile'}), 500
        try:
            cursor = conn.cursor()
            cursor.execute("SELECT key, value FROM system_settings WHERE key LIKE 'backup.%'")
            settings = dict(cursor.fetchall())
            return jsonify({'success': True, 'schedule': settings})
        finally:
            conn.close()
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/backup/config', methods=['GET', 'POST'])
@require_auth()
@require_permission('all')
def api_backup_config():
    if request.method == 'GET':
        return backup_module.load_config()
    else:
        config = request.get_json()
        return backup_module.save_config(config)

# ===============================
# API ENDPOINTS CONFIGURAZIONI SISTEMA
# ===============================

@app.route('/api/system/config/save', methods=['POST'])
@require_auth()
@require_permission('all')
def api_save_system_config():
    """Salva configurazioni sistema"""
    try:
        data = request.get_json()
        conn = get_db_connection()
        if not conn:
            return jsonify({'success': False, 'error': 'Database non disponibile'}), 500
        try:
            cursor = conn.cursor()
            for section, values in data.items():
                for key, value in values.items():
                    # Forza il prefisso 'sistema.' per tutte le chiavi di sistema
                    if not section.startswith('sistema'):
                        config_key = f"sistema.{key}"
                    else:
                        config_key = f"{section}.{key}"
                    if isinstance(value, (dict, list)):
                        import json
                        json_value = json.dumps(value)
                    else:
                        json_value = str(value)
                    cursor.execute('''
                        INSERT OR REPLACE INTO system_settings (key, value, updated_at)
                        VALUES (?, ?, CURRENT_TIMESTAMP)
                    ''', (config_key, json_value))
            conn.commit()
            # Log evento
            cursor.execute('''
                INSERT INTO eventi_sistema (tipo_evento, livello, messaggio, componente)
                VALUES (?, ?, ?, ?)
            ''', ('CONFIG_UPDATE', 'INFO', 'Configurazioni sistema aggiornate', 'SYSTEM'))
            conn.commit()
            return jsonify({
                'success': True, 
                'message': 'Configurazioni salvate con successo'
            })
        finally:
            conn.close()
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

# 3. AGGIUNGI il nuovo endpoint per il reset (dopo api_system_restart, circa riga 620)
@app.route('/api/system/config/reset', methods=['POST'])
@require_auth()
@require_permission('all')
def api_reset_system_config():
    """Ripristina configurazioni ai valori default"""
    try:
        # Ottieni configurazioni default
        conn = get_db_connection()
        if not conn:
            return jsonify({'success': False, 'error': 'Database non disponibile'}), 500
        
        try:
            cursor = conn.cursor()
            
            # Crea tabella se non esiste
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS system_settings (
                    key TEXT PRIMARY KEY,
                    value TEXT NOT NULL,
                    updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            ''')
            
            # Elimina tutte le configurazioni personalizzate
            cursor.execute('DELETE FROM system_settings')
            
            # Inserisci valori default
            default_values = {
                'sistema.nome_installazione': 'Isola Ecologica RAEE - Rende',
                'sistema.versione': '1.0.0',
                'sistema.ambiente': 'production',
                'sistema.debug_mode': 'false',
                'sistema.porta_web': '5000',
                'sistema.timeout_sessione': '1800',
                
                'hardware.lettore_porta': '/dev/ttyACM0',
                'hardware.relay_porta': '/dev/ttyUSB0',
                'hardware.relay_baudrate': '19200',
                'hardware.gate_duration': '8.0',
                
                'sicurezza.max_tentativi_login': '5',
                'sicurezza.durata_blocco_minuti': '15',
                'sicurezza.rotazione_password_giorni': '90',
                'sicurezza.log_audit_abilitato': 'true',
                
                'email.smtp_server': '',
                'email.smtp_porta': '587',
                'email.mittente': '',
                'email.report_automatici': 'false',
                'email.frequenza_report': 'weekly'
            }
            
            for key, value in default_values.items():
                cursor.execute('''
                    INSERT INTO system_settings (key, value, updated_at)
                    VALUES (?, ?, CURRENT_TIMESTAMP)
                ''', (key, value))
            
            conn.commit()
            
            # Log evento
            cursor.execute('''
                INSERT INTO eventi_sistema (tipo_evento, livello, messaggio, componente)
                VALUES (?, ?, ?, ?)
            ''', ('CONFIG_RESET', 'WARNING', 'Configurazioni ripristinate ai valori default', 'SYSTEM'))
            
            conn.commit();
            
            return jsonify({
                'success': True,
                'message': 'Configurazioni ripristinate ai valori default'
            })
            
        finally:
            conn.close()
            
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/system/restart', methods=['POST'])
@require_auth()
@require_permission('all')
def api_system_restart():
    """Riavvia realmente il sistema tramite systemd"""
    try:
        import subprocess
        # Riavvia il servizio systemd (modifica il nome se necessario)
        subprocess.Popen(['sudo', 'systemctl', 'restart', 'access-control-system'])
        return jsonify({
            'success': True,
            'message': 'Sistema in riavvio...'
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/restart-service', methods=['POST'])
@require_auth()
@require_permission('all')
def api_restart_service():
    """Riavvia il servizio web API"""
    try:
        import os
        import signal
        import threading
        import time
        
        logger.info("Richiesta riavvio servizio da Dashboard Debug...")
        
        # Funzione per riavviare il servizio dopo un breve delay
        def delayed_restart():
            time.sleep(1)
            try:
                # Prova prima a riavviare con systemctl se disponibile
                import subprocess
                result = subprocess.run(['systemctl', 'restart', 'access-control-web'],
                                      capture_output=True, text=True, timeout=5)
                if result.returncode == 0:
                    logger.info("Servizio riavviato con systemctl")
                    return
            except Exception:
                pass
            
            # Altrimenti usa exec per riavviare il processo Python
            logger.info("Riavvio processo Python...")
            os.execv(sys.executable, [sys.executable] + sys.argv)
        
        # Avvia il restart in un thread separato
        restart_thread = threading.Thread(target=delayed_restart, daemon=True)
        restart_thread.start()
        
        return jsonify({
            'success': True,
            'message': 'Servizio in riavvio. Attendere 3-5 secondi...'
        })
    except Exception as e:
        logger.error(f"Errore riavvio servizio: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/system-logs')
@require_auth()
@require_permission('all')
def api_system_logs():
    """Restituisce gli ultimi log del sistema"""
    try:
        import subprocess
        from collections import deque
        
        # Permetti di specificare il numero di righe da recuperare (default 30, max 2000)
        num_lines = request.args.get('lines', 30, type=int)
        num_lines = min(num_lines, 2000)  # Limita a 2000 per sicurezza
        
        # Prova a leggere i log dal journal di systemd
        try:
            result = subprocess.run(
                ['journalctl', '-u', 'access-control-web', '-n', str(num_lines + 20), '--no-pager'],
                capture_output=True,
                text=True,
                timeout=2
            )
            if result.returncode == 0 and result.stdout:
                logs = result.stdout.strip().split('\n')[-num_lines:]  # Ultimi N righe
                return jsonify({'success': True, 'logs': logs})
        except:
            pass
        
        # Altrimenti usa i log del logger Python
        logs = []
        
        # Aggiungi alcuni log di stato
        logs.append(f"[{datetime.now().strftime('%H:%M:%S')}] INFO - Sistema operativo")
        logs.append(f"[{datetime.now().strftime('%H:%M:%S')}] INFO - Dashboard accessibile su http://192.168.1.236:5000")
        
        # Controlla stato hardware
        if 'card_reader_running' in globals() and card_reader_running:
            logs.append(f"[{datetime.now().strftime('%H:%M:%S')}] INFO - ✅ Lettore tessere attivo")
        else:
            logs.append(f"[{datetime.now().strftime('%H:%M:%S')}] WARNING - ⚠️ Lettore tessere non attivo")
            
        return jsonify({'success': True, 'logs': logs})
        
    except Exception as e:
        logger.error(f"Errore recupero log: {e}")
        return jsonify({'success': False, 'error': str(e), 'logs': []})

@app.route('/api/system-status')
@require_auth()
@require_permission('all')
def api_system_status():
    """Restituisce lo stato del sistema"""
    try:
        status = {
            'service_running': True,  # Se risponde, è attivo
            'reader_connected': False,
            'reader_type': None,
            'relay_connected': False,
            'database_ok': False
        }
        
        # Controlla lettore tessere
        # Prima verifica se c'è un lettore configurato
        try:
            # Controlla se CRT-285 è configurato
            config_path = '/opt/access_control/config/device_assignments.json'
            if os.path.exists(config_path):
                with open(config_path, 'r') as f:
                    config = json.load(f)
                    if 'assignments' in config and 'card_reader' in config['assignments']:
                        reader_info = config['assignments']['card_reader']
                        if '23d8:0285' in reader_info.get('device_key', ''):
                            status['reader_connected'] = True
                            status['reader_type'] = 'CRT-285'
                        elif 'card_reader_running' in globals() and card_reader_running:
                            status['reader_connected'] = True
                            status['reader_type'] = reader_info.get('device_name', 'Lettore generico')
        except Exception as e:
            logger.debug(f"Errore verifica lettore: {e}")
        
        # Fallback al controllo della variabile globale
        if not status['reader_connected'] and 'card_reader_running' in globals() and card_reader_running:
            status['reader_connected'] = True
            status['reader_type'] = 'Omnikey'
        
        # Controlla database
        try:
            conn = get_db_connection()
            if conn:
                status['database_ok'] = True
                conn.close()
        except:
            pass
            
        # Controlla relay
        try:
            if os.path.exists('/dev/ttyACM0'):
                status['relay_connected'] = True
        except:
            pass
            
        return jsonify({
            'success': True,
            **status
        })
        
    except Exception as e:
        logger.error(f"Errore stato sistema: {e}")
        return jsonify({
            'success': False,
            'error': str(e),
            'service_running': False
        })

# ===============================
# API ENDPOINTS CONFIGURAZIONI - USANDO CONFIG ESISTENTE
# ===============================



@app.route('/api/email/config')
@require_auth()
def api_get_email_config():
    """Ottieni configurazione email"""
    try:
        # Carica da database temporaneo
        conn = get_db_connection()
        if not conn:
            return jsonify({'success': False, 'error': 'Database non disponibile'}), 500
        
        try:
            cursor = conn.cursor()
            cursor.execute("SELECT key, value FROM system_settings WHERE key LIKE 'email.%'")
            settings = cursor.fetchall()
            
            config = {
                'smtp_server': '',
                'smtp_port': 587,
                'mittente': '',
                'frequenza_report': 'weekly',
                'orario_invio': '08:00',
                'includi_statistiche': True,
                'includi_errori': True
            }
            
            # Popola con valori salvati
            for key, value in settings:
                setting_name = key.replace('email.', '')
                if setting_name in config:
                    try:
                        # Prova a convertire numeri e boolean
                        if value.lower() in ('true', 'false'):
                            config[setting_name] = value.lower() == 'true'
                        elif value.isdigit():
                            config[setting_name] = int(value)
                        else:
                            config[setting_name] = value
                    except:
                        config[setting_name] = value
            
            return jsonify({'success': True, 'config': config})
            
        finally:
            conn.close()
            
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/email/save-smtp', methods=['POST'])
@require_auth()
@require_permission('all')
def api_save_smtp_config():
    """Salva configurazione SMTP"""
    try:
        data = request.get_json()
        
        conn = get_db_connection()
        if not conn:
            return jsonify({'success': False, 'error': 'Database non disponibile'}), 500
        
        try:
            cursor = conn.cursor()
            
            # Crea tabella se non esiste
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS system_settings (
                    key TEXT PRIMARY KEY,
                    value TEXT NOT NULL,
                    updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            ''')
            
            # Salva configurazioni email
            email_settings = {
                'smtp_server': data.get('smtp_server', ''),
                'smtp_port': str(data.get('smtp_port', 587)),
                'mittente': data.get('mittente', '')
            }
            
            for key, value in email_settings.items():
                cursor.execute('''
                    INSERT OR REPLACE INTO system_settings (key, value, updated_at)
                    VALUES (?, ?, CURRENT_TIMESTAMP)
                ''', (f'email.{key}', value))
            
            # Non salvare password per sicurezza
            if data.get('password'):
                cursor.execute('''
                    INSERT OR REPLACE INTO system_settings (key, value, updated_at)
                    VALUES (?, ?, CURRENT_TIMESTAMP)
                ''', ('email.password_set', 'true'))
            
            conn.commit()
            
            return jsonify({'success': True, 'message': 'Configurazione SMTP salvata'})
            
        finally:
            conn.close()
            
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/devices/config')
@require_auth()
def api_get_devices_config():
    """Ottieni configurazione dispositivi"""
    try:
        # Carica da database temporaneo
        conn = get_db_connection()
        if not conn:
            return jsonify({'success': False, 'error': 'Database non disponibile'}), 500
        
        try:
            cursor = conn.cursor()
            cursor.execute("SELECT key, value FROM system_settings WHERE key LIKE 'dispositivi.%'")
            settings = cursor.fetchall()
            
            config = {
                'reader': {
                    'tipo': 'HID Omnikey 5427CK',
                    'porta': '/dev/ttyACM0',
                    'timeout': 30,
                    'retry_tentativi': 3
                },
                'relay': {
                    'porta': '/dev/ttyUSB0',
                    'baud_rate': 19200,
                    'modulo_id': 8
                }
            }
            
            # Popola con valori salvati
            for key, value in settings:
                parts = key.split('.')
                if len(parts) >= 3:
                    device = parts[1]  # lettore_cf o rele
                    setting = parts[2]
                    
                    if device == 'lettore_cf' and setting in config['reader']:
                        try:
                            if setting in ['timeout', 'retry_tentativi']:
                                config['reader'][setting] = int(value)
                            else:
                                config['reader'][setting] = value
                        except:
                            config['reader'][setting] = value
                    elif device == 'rele' and setting in config['relay']:
                        try:
                            if setting in ['baud_rate', 'modulo_id']:
                                config['relay'][setting] = int(value)
                            else:
                                config['relay'][setting] = value
                        except:
                            config['relay'][setting] = value
            
            return jsonify({'success': True, **config})
            
        finally:
            conn.close()
            
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/devices/config/reader', methods=['POST'])
@require_auth()
@require_permission('all')
def api_save_reader_config():
    """Salva configurazione lettore nel database"""
    try:
        data = request.get_json()
        if not data.get('tipo') or not data.get('porta'):
            return jsonify({'success': False, 'error': 'Configurazione incompleta'}), 400

        conn = get_db_connection()
        if not conn:
            return jsonify({'success': False, 'error': 'Database non disponibile'}), 500

        try:
            cursor = conn.cursor()
            settings = {
                'dispositivi.lettore_cf.tipo': data['tipo'],
                'dispositivi.lettore_cf.porta': data['porta'],
                'dispositivi.lettore_cf.timeout': str(data.get('timeout', 30)),
                'dispositivi.lettore_cf.retry_tentativi': str(data.get('retry_tentativi', 3))
            }
            for key, value in settings.items():
                cursor.execute('''
                    INSERT OR REPLACE INTO system_settings (key, value, updated_at)
                    VALUES (?, ?, CURRENT_TIMESTAMP)
                ''', (key, value))
            conn.commit()
            return jsonify({'success': True, 'message': 'Configurazione lettore salvata'})
        finally:
            conn.close()
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/devices/config/relay', methods=['POST'])
@require_auth()
@require_permission('all')
def api_save_relay_config():
    """Salva configurazione relè nel database"""
    try:
        data = request.get_json()
        if not data.get('porta') or not data.get('baud_rate'):
            return jsonify({'success': False, 'error': 'Configurazione incompleta'}), 400

        conn = get_db_connection()
        if not conn:
            return jsonify({'success': False, 'error': 'Database non disponibile'}), 500

        try:
            cursor = conn.cursor()
            settings = {
                'dispositivi.rele.porta': data['porta'],
                'dispositivi.rele.baud_rate': str(data['baud_rate']),
                'dispositivi.rele.modulo_id': str(data.get('modulo_id', 8))
            }
            for key, value in settings.items():
                cursor.execute('''
                    INSERT OR REPLACE INTO system_settings (key, value, updated_at)
                    VALUES (?, ?, CURRENT_TIMESTAMP)
                ''', (key, value))
            # Salva mapping se presente
            if 'mapping' in data:
                import json
                cursor.execute('''
                    INSERT OR REPLACE INTO system_settings (key, value, updated_at)
                    VALUES (?, ?, CURRENT_TIMESTAMP)
                ''', ('dispositivi.rele.mapping', json.dumps(data['mapping'])))
            conn.commit()
            return jsonify({'success': True, 'message': 'Configurazione relè salvata'})
        finally:
            conn.close()
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/system/config')
@require_auth()
def api_get_system_config():
    """Restituisce tutte le configurazioni di sistema dal database"""
    try:
        conn = get_db_connection()
        if not conn:
            return jsonify({'success': False, 'error': 'Database non disponibile'}), 500

        try:
            cursor = conn.cursor()
            # Prendi tutte le chiavi di sistema, hardware, sicurezza, email
            cursor.execute("""
                SELECT key, value FROM system_settings
                WHERE key LIKE 'sistema.%'
                   OR key LIKE 'hardware.%'
                   OR key LIKE 'sicurezza.%'
                   OR key LIKE 'email.%'
            """)
            settings = cursor.fetchall()
            config = {}
            for key, value in settings:
                section, name = key.split('.', 1)
                if section not in config:
                    config[section] = {}
                config[section][name] = value
            return jsonify({'success': True, 'config': config})
        finally:
            conn.close()
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/relay-config', methods=['GET'])
@require_auth()
def get_relay_config():
    """Restituisce la configurazione dei relè"""
    try:
        conn = get_db_connection()
        if not conn:
            return jsonify({'success': False, 'error': 'Database non disponibile'}), 500
        
        try:
            cursor = conn.cursor()
            cursor.execute("""
                SELECT relay_number, description, valid_action, valid_duration, 
                       invalid_action, invalid_duration 
                FROM relay_config
                ORDER BY relay_number
            """)
            rows = cursor.fetchall()
            
            config = {}
            for row in rows:
                config[f'relay_{row[0]}'] = {
                    'description': row[1],
                    'valid_action': row[2],
                    'valid_duration': row[3],
                    'invalid_action': row[4],
                    'invalid_duration': row[5]
                }
            
            return jsonify(config)
        finally:
            conn.close()
    except Exception as e:
        logger.error(f"Errore recupero configurazione relè: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/relay-config', methods=['POST'])
@require_auth()
def save_relay_config():
    """Salva la configurazione dei relè"""
    try:
        data = request.json
        conn = get_db_connection()
        if not conn:
            return jsonify({'success': False, 'error': 'Database non disponibile'}), 500
        
        try:
            cursor = conn.cursor()
            
            # Aggiorna configurazione per ogni relè
            for relay_key, config in data.items():
                if relay_key.startswith('relay_'):
                    relay_num = int(relay_key.split('_')[1])
                    
                    cursor.execute("""
                        INSERT OR REPLACE INTO relay_config 
                        (relay_number, description, valid_action, valid_duration, 
                         invalid_action, invalid_duration, updated_at, updated_by)
                        VALUES (?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP, ?)
                    """, (
                        relay_num,
                        config.get('description', f'Relè {relay_num}'),
                        config.get('valid_action', 'OFF'),
                        float(config.get('valid_duration', 0)),
                        config.get('invalid_action', 'OFF'),
                        float(config.get('invalid_duration', 0)),
                        session.get('username', 'sistema')
                    ))
            
            conn.commit()
            logger.info(f"Configurazione relè salvata da {session.get('username', 'sistema')}")
            return jsonify({'success': True})
        finally:
            conn.close()
    except Exception as e:
        logger.error(f"Errore salvataggio configurazione relè: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/server-time')
@require_auth()
def api_get_server_time():
    """API per ottenere l'ora del server"""
    from datetime import datetime
    import pytz
    
    # Recupera timezone configurato
    timezone_name = 'Europe/Rome'  # Default
    try:
        conn = get_db_connection()
        if conn:
            cursor = conn.cursor()
            cursor.execute("SELECT value FROM system_settings WHERE key = 'sistema.timezone'")
            result = cursor.fetchone()
            if result:
                timezone_name = result[0]
            conn.close()
    except Exception as e:
        logger.error(f"Errore recupero timezone: {e}")
    
    # Usa il timezone configurato
    tz = pytz.timezone(timezone_name)
    now = datetime.now(tz)
    
    return jsonify({
        'time': now.strftime('%H:%M:%S'),
        'date': now.strftime('%d/%m/%Y'),
        'weekday': now.strftime('%A'),
        'timestamp': now.isoformat(),
        'timezone': timezone_name
    })

@app.route('/api/log-accessi')
@require_auth()
def api_get_log_accessi():
    """API per recuperare i log accessi con filtri e paginazione"""
    from datetime import datetime
    import pytz
    
    try:
        # Recupera timezone configurato
        timezone_name = 'Europe/Rome'  # Default
        try:
            conn_tz = get_db_connection()
            if conn_tz:
                cursor_tz = conn_tz.cursor()
                cursor_tz.execute("SELECT value FROM system_settings WHERE key = 'sistema.timezone'")
                result_tz = cursor_tz.fetchone()
                if result_tz:
                    timezone_name = result_tz[0]
                conn_tz.close()
        except Exception as e:
            logger.error(f"Errore recupero timezone: {e}")
        
        tz = pytz.timezone(timezone_name)
        utc = pytz.utc
        
        # Parametri filtri
        page = int(request.args.get('page', 1))
        periodo = request.args.get('periodo', 'mese')
        data_inizio = request.args.get('data_inizio', '')
        data_fine = request.args.get('data_fine', '')
        tipo_accesso = request.args.get('tipo', '')
        codice_fiscale = request.args.get('codice_fiscale', '')
        
        # Configurazione paginazione
        records_per_page = 50
        offset = (page - 1) * records_per_page
        
        conn = get_db_connection()
        if not conn:
            return jsonify({'error': 'Database non disponibile'}), 500
        
        cursor = conn.cursor()
        
        # Costruisci query con filtri - join con utenti_autorizzati per il nome
        query = """
            SELECT l.*, u.nome as nome_from_users 
            FROM log_accessi l
            LEFT JOIN utenti_autorizzati u ON l.codice_fiscale = u.codice_fiscale
            WHERE 1=1
        """
        params = []
        
        # Filtro periodo
        if periodo == 'oggi':
            query += " AND DATE(timestamp) = DATE('now')"
        elif periodo == 'settimana':
            query += " AND DATE(timestamp) >= DATE('now', '-7 days')"
        elif periodo == 'mese':
            query += " AND strftime('%Y-%m', timestamp) = strftime('%Y-%m', 'now')"
        elif periodo == 'custom' and data_inizio and data_fine:
            query += " AND DATE(timestamp) BETWEEN ? AND ?"
            params.extend([data_inizio, data_fine])
        
        # Filtro tipo accesso
        if tipo_accesso:
            query += " AND tipo_accesso = ?"
            params.append(tipo_accesso)
        
        # Filtro codice fiscale
        if codice_fiscale:
            query += " AND codice_fiscale LIKE ?"
            params.append(f'%{codice_fiscale}%')
        
        # Conta totale record con filtri
        count_query = query.replace("SELECT l.*, u.nome as nome_from_users", "SELECT COUNT(*)")
        cursor.execute(count_query, params)
        total_records = cursor.fetchone()[0]
        total_pages = (total_records + records_per_page - 1) // records_per_page
        
        # Query con paginazione
        query += " ORDER BY timestamp DESC LIMIT ? OFFSET ?"
        params.extend([records_per_page, offset])
        
        cursor.execute(query, params)
        logs = []
        for row in cursor.fetchall():
            # Converti timestamp da UTC a timezone configurato
            timestamp_str = row[1]
            try:
                # Parse timestamp from database (assumed to be UTC)
                dt_utc = datetime.strptime(timestamp_str, '%Y-%m-%d %H:%M:%S')
                dt_utc = utc.localize(dt_utc)
                # Convert to configured timezone
                dt_local = dt_utc.astimezone(tz)
                timestamp_converted = dt_local.strftime('%Y-%m-%d %H:%M:%S')
            except Exception as e:
                logger.error(f"Errore conversione timestamp: {e}")
                timestamp_converted = timestamp_str
            
            # Map columns correctly
            log_entry = {
                'id': row[0],
                'timestamp': timestamp_converted,
                'codice_fiscale': row[2],
                'autorizzato': row[3],
                'durata_elaborazione': float(row[4]) if row[4] else 0.0,
                'terminale_id': row[7] if len(row) > 7 else None,
                'tipo_accesso': row[14] if len(row) > 14 else None,
                'motivo_rifiuto': row[15] if len(row) > 15 else None,
                'nome_utente': row[16] if len(row) > 16 else None
            }
            
            # Use joined nome_utente from utenti_autorizzati if available
            if len(row) > 17 and row[17]:
                log_entry['nome_utente'] = row[17]
            
            # Se nome_utente ancora vuoto, mostra '-'
            if not log_entry['nome_utente']:
                log_entry['nome_utente'] = '-'
                
            # Format tipo_accesso properly if missing or 'NO'
            if not log_entry['tipo_accesso'] or log_entry['tipo_accesso'] == 'NO':
                if log_entry['autorizzato']:
                    log_entry['tipo_accesso'] = 'AUTORIZZATO'
                elif log_entry['motivo_rifiuto']:
                    motivo_lower = str(log_entry['motivo_rifiuto']).lower()
                    if 'orario' in motivo_lower:
                        log_entry['tipo_accesso'] = 'FUORI_ORARIO'
                    elif 'limite' in motivo_lower:
                        log_entry['tipo_accesso'] = 'LIMITE_SUPERATO'
                    elif 'disattivato' in motivo_lower:
                        log_entry['tipo_accesso'] = 'UTENTE_DISATTIVATO'
                    elif 'non trovato' in motivo_lower:
                        log_entry['tipo_accesso'] = 'UTENTE_NON_TROVATO'
                    else:
                        log_entry['tipo_accesso'] = 'NEGATO'
                else:
                    log_entry['tipo_accesso'] = 'NEGATO'
            
            # Ensure durata_elaborazione is in seconds (convert from microseconds if needed)
            if log_entry['durata_elaborazione'] and log_entry['durata_elaborazione'] < 0.01:
                log_entry['durata_elaborazione'] = log_entry['durata_elaborazione'] * 1000  # Convert to ms
            
            logs.append(log_entry)
        
        # Calcola statistiche
        stats_query = """
            SELECT 
                SUM(CASE WHEN autorizzato = 1 THEN 1 ELSE 0 END) as autorizzati,
                SUM(CASE WHEN autorizzato = 0 THEN 1 ELSE 0 END) as negati,
                SUM(CASE WHEN tipo_accesso = 'FUORI_ORARIO' THEN 1 ELSE 0 END) as fuori_orario,
                SUM(CASE WHEN DATE(timestamp) = DATE('now') THEN 1 ELSE 0 END) as oggi
            FROM log_accessi
        """
        cursor.execute(stats_query)
        stats = cursor.fetchone()
        
        conn.close()
        
        return jsonify({
            'logs': logs,
            'total_records': total_records,
            'total_pages': total_pages,
            'current_page': page,
            'statistics': {
                'autorizzati': stats[0] or 0,
                'negati': stats[1] or 0,
                'fuori_orario': stats[2] or 0,
                'oggi': stats[3] or 0
            }
        })
    except Exception as e:
        logger.error(f"Errore recupero log accessi: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/log-accessi/export')
@require_auth()
def api_export_log_accessi():
    """Export log accessi in Excel/CSV/PDF"""
    try:
        import csv
        import io
        from datetime import datetime
        import pytz
        
        format_type = request.args.get('format', 'csv')
        
        # Recupera timezone configurato
        timezone_name = 'Europe/Rome'  # Default
        try:
            conn_tz = get_db_connection()
            if conn_tz:
                cursor_tz = conn_tz.cursor()
                cursor_tz.execute("SELECT value FROM system_settings WHERE key = 'sistema.timezone'")
                result_tz = cursor_tz.fetchone()
                if result_tz:
                    timezone_name = result_tz[0]
                conn_tz.close()
        except Exception as e:
            logger.error(f"Errore recupero timezone: {e}")
        
        tz = pytz.timezone(timezone_name)
        utc = pytz.utc
        
        # Applica gli stessi filtri della visualizzazione
        periodo = request.args.get('periodo', 'mese')
        data_inizio = request.args.get('data_inizio', '')
        data_fine = request.args.get('data_fine', '')
        tipo_accesso = request.args.get('tipo', '')
        codice_fiscale = request.args.get('codice_fiscale', '')
        
        conn = get_db_connection()
        if not conn:
            return jsonify({'error': 'Database non disponibile'}), 500
        
        cursor = conn.cursor()
        
        # Costruisci query con filtri (come sopra)
        query = """
            SELECT l.*, u.nome as nome_from_users 
            FROM log_accessi l
            LEFT JOIN utenti_autorizzati u ON l.codice_fiscale = u.codice_fiscale
            WHERE 1=1
        """
        params = []
        
        if periodo == 'oggi':
            query += " AND DATE(timestamp) = DATE('now')"
        elif periodo == 'settimana':
            query += " AND DATE(timestamp) >= DATE('now', '-7 days')"
        elif periodo == 'mese':
            query += " AND strftime('%Y-%m', timestamp) = strftime('%Y-%m', 'now')"
        elif periodo == 'custom' and data_inizio and data_fine:
            query += " AND DATE(timestamp) BETWEEN ? AND ?"
            params.extend([data_inizio, data_fine])
        
        if tipo_accesso:
            query += " AND tipo_accesso = ?"
            params.append(tipo_accesso)
        
        if codice_fiscale:
            query += " AND codice_fiscale LIKE ?"
            params.append(f'%{codice_fiscale}%')
        
        query += " ORDER BY timestamp DESC"
        
        if format_type == 'csv':
            cursor.execute(query, params)
            # Crea CSV
            output = io.StringIO()
            writer = csv.writer(output)
            writer.writerow(['Data/Ora', 'Codice Fiscale', 'Nome Utente', 'Autorizzato', 'Tipo Accesso', 'Motivo Rifiuto', 'Terminale', 'Durata (ms)'])
            
            for row in cursor.fetchall():
                # Converti timestamp da UTC a timezone configurato
                timestamp_str = row[1]
                try:
                    dt_utc = datetime.strptime(timestamp_str, '%Y-%m-%d %H:%M:%S')
                    dt_utc = utc.localize(dt_utc)
                    dt_local = dt_utc.astimezone(tz)
                    timestamp_converted = dt_local.strftime('%Y-%m-%d %H:%M:%S')
                except:
                    timestamp_converted = timestamp_str
                
                writer.writerow([
                    timestamp_converted,  # timestamp convertito
                    row[2],  # codice_fiscale  
                    row[17] if len(row) > 17 else row[16] if len(row) > 16 else '-',  # nome_utente dalla join
                    'Si' if row[3] else 'No',  # autorizzato
                    row[14] if len(row) > 14 else '',  # tipo_accesso
                    row[15] if len(row) > 15 else '',  # motivo_rifiuto
                    row[7] or '',  # terminale_id
                    f"{float(row[4]):.2f}" if row[4] else '0.00'  # durata_elaborazione
                ])
            
            conn.close()
            
            # Prepara response
            output.seek(0)
            response = make_response(output.getvalue())
            response.headers['Content-Disposition'] = f'attachment; filename=log_accessi_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'
            response.headers['Content-Type'] = 'text/csv'
            return response
            
        elif format_type == 'excel':
            cursor.execute(query, params)
            # Per Excel useremo openpyxl se disponibile
            try:
                from openpyxl import Workbook
                from openpyxl.styles import Font, PatternFill
                
                wb = Workbook()
                ws = wb.active
                ws.title = "Log Accessi"
                
                # Header
                headers = ['Data/Ora', 'Codice Fiscale', 'Nome Utente', 'Autorizzato', 'Tipo Accesso', 'Motivo Rifiuto', 'Terminale', 'Durata (ms)']
                ws.append(headers)
                
                # Stile header
                for cell in ws[1]:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                
                # Dati
                for row in cursor.fetchall():
                    # Converti timestamp da UTC a timezone configurato
                    timestamp_str = row[1]
                    try:
                        dt_utc = datetime.strptime(timestamp_str, '%Y-%m-%d %H:%M:%S')
                        dt_utc = utc.localize(dt_utc)
                        dt_local = dt_utc.astimezone(tz)
                        timestamp_converted = dt_local.strftime('%Y-%m-%d %H:%M:%S')
                    except:
                        timestamp_converted = timestamp_str
                    
                    ws.append([
                        timestamp_converted,
                        row[2],
                        row[17] if len(row) > 17 else row[16] if len(row) > 16 else '-',
                        'Si' if row[3] else 'No',
                        row[14] if len(row) > 14 else '',
                        row[15] if len(row) > 15 else '',
                        row[7] or '',
                        f"{float(row[4]):.2f}" if row[4] else '0.00'
                    ])
                
                conn.close()
                
                # Salva in BytesIO
                excel_file = io.BytesIO()
                wb.save(excel_file)
                excel_file.seek(0)
                
                response = make_response(excel_file.getvalue())
                response.headers['Content-Disposition'] = f'attachment; filename=log_accessi_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
                response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                return response
                
            except ImportError:
                return jsonify({'error': 'openpyxl non installato'}), 500
        
        elif format_type == 'pdf':
            cursor.execute(query, params)
            # Per PDF useremo reportlab
            try:
                from reportlab.lib import colors
                from reportlab.lib.pagesizes import A4, landscape
                from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
                from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
                from reportlab.lib.units import inch
                
                # Crea buffer per PDF
                pdf_file = io.BytesIO()
                doc = SimpleDocTemplate(pdf_file, pagesize=landscape(A4), rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=18)
                
                # Container per gli elementi
                elements = []
                
                # Stili
                styles = getSampleStyleSheet()
                title_style = ParagraphStyle(
                    'CustomTitle',
                    parent=styles['Heading1'],
                    fontSize=24,
                    textColor=colors.HexColor('#2c3e50'),
                    spaceAfter=30,
                    alignment=1  # Center
                )
                
                # Titolo
                title = Paragraph("Log Accessi - Sistema Controllo Accessi", title_style)
                elements.append(title)
                
                # Prepara i dati per la tabella
                data = [['Data/Ora', 'Codice Fiscale', 'Nome Utente', 'Autorizzato', 'Tipo Accesso', 'Terminale']]
                
                rows = cursor.fetchall()
                if not rows:
                    # Se non ci sono dati, aggiungi una riga vuota
                    data.append(['Nessun dato disponibile', '', '', '', '', ''])
                else:
                    for row in rows:
                        # Converti timestamp da UTC a timezone configurato
                        timestamp_str = row[1] if row[1] else ''
                        try:
                            if timestamp_str:
                                dt_utc = datetime.strptime(timestamp_str, '%Y-%m-%d %H:%M:%S')
                                dt_utc = utc.localize(dt_utc)
                                dt_local = dt_utc.astimezone(tz)
                                timestamp_converted = dt_local.strftime('%d/%m/%Y %H:%M')
                            else:
                                timestamp_converted = ''
                        except:
                            timestamp_converted = timestamp_str
                        
                        nome_utente = '-'
                        if len(row) > 17 and row[17]:
                            nome_utente = str(row[17])[:20]
                        elif len(row) > 16 and row[16]:
                            nome_utente = str(row[16])[:20]
                        
                        data.append([
                            timestamp_converted,
                            str(row[2])[:16] if row[2] else '',  # codice fiscale
                            nome_utente,  # nome utente
                            'Si' if row[3] else 'No',
                            str(row[14])[:15] if len(row) > 14 and row[14] else '',  # tipo_accesso
                            str(row[7])[:15] if row[7] else ''  # terminale
                        ])
                
                # Crea tabella
                table = Table(data)
                
                # Stile tabella
                table.setStyle(TableStyle([
                    # Header
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3498db')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 12),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    
                    # Righe dati
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                    ('FONTSIZE', (0, 1), (-1, -1), 9),
                    ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
                    
                    # Colori alternati per le righe
                    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#ecf0f1')]),
                ]))
                
                elements.append(table)
                
                # Genera PDF
                doc.build(elements)
                conn.close()
                
                pdf_file.seek(0)
                response = make_response(pdf_file.getvalue())
                response.headers['Content-Disposition'] = f'attachment; filename=log_accessi_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
                response.headers['Content-Type'] = 'application/pdf'
                return response
                
            except ImportError:
                return jsonify({'error': 'reportlab non installato'}), 500
        
        else:
            return jsonify({'error': 'Formato non supportato'}), 400
            
    except Exception as e:
        logger.error(f"Errore export log accessi: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/relay-config/test', methods=['POST'])
@require_auth()
def test_relay_config():
    """Testa la configurazione dei relè (simula tessera valida)"""
    try:
        # Crea una nuova connessione al controller (come fa test_gate)
        from hardware.usb_rly08_controller import USBRLY08Controller
        controller = USBRLY08Controller()
        
        if not controller.connect():
            return jsonify({'success': False, 'error': 'Impossibile connettersi al controller relè'}), 500
        
        try:
            conn = get_db_connection()
            if not conn:
                controller.disconnect()
                return jsonify({'success': False, 'error': 'Database non disponibile'}), 500
            
            try:
                cursor = conn.cursor()
                cursor.execute("""
                    SELECT relay_number, description, valid_action, valid_duration 
                    FROM relay_config
                    WHERE valid_action != 'OFF'
                    ORDER BY relay_number
                """)
                configs = cursor.fetchall()
                
                # Applica configurazione per tessera valida
                for relay_num, description, action, duration in configs:
                    logger.info(f"Test relè {relay_num} ({description}): {action} per {duration}s")
                    
                    if action == 'ON':
                        # Accendi relè
                        controller._send_command(100 + relay_num)
                        if duration > 0:
                            # Programma spegnimento
                            threading.Timer(duration, lambda rn=relay_num, ctrl=controller: ctrl._send_command(110 + rn)).start()
                            
                    elif action == 'PULSE':
                        # Impulso: accendi e spegni dopo durata
                        controller._send_command(100 + relay_num)
                        if duration > 0:
                            threading.Timer(duration, lambda rn=relay_num, ctrl=controller: ctrl._send_command(110 + rn)).start()
                        else:
                            # Impulso breve default 500ms
                            threading.Timer(0.5, lambda rn=relay_num, ctrl=controller: ctrl._send_command(110 + rn)).start()
                
                # Disconnetti dopo un delay per permettere ai timer di eseguire
                threading.Timer(10.0, controller.disconnect).start()
                
                return jsonify({'success': True, 'message': 'Test configurazione avviato'})
            finally:
                conn.close()
        except Exception as e:
            controller.disconnect()
            raise e
            
    except Exception as e:
        logger.error(f"Errore test configurazione relè: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

BACKUP_DIR = os.path.join(project_root, 'backups')
if not os.path.exists(BACKUP_DIR):
    os.makedirs(BACKUP_DIR)

def get_db_connection():
    """Connessione database come database_manager.py"""
    try:
        # Crea la directory se non esiste
        db_dir = os.path.dirname(DB_PATH)
        if not os.path.exists(db_dir):
            os.makedirs(db_dir)
            
        # Crea il database se non esiste
        if not os.path.exists(DB_PATH):
            conn = sqlite3.connect(DB_PATH)
            conn.row_factory = sqlite3.Row
            print(f"Database creato in: {DB_PATH}")
            return conn
            
        conn = sqlite3.connect(DB_PATH)
        conn.row_factory = sqlite3.Row
        return conn
    except Exception as e:
        print(f"Errore connessione DB: {e}")
        return None

def ensure_system_settings_table():
    conn = get_db_connection()
    if conn:
        try:
            cursor = conn.cursor()
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS system_settings (
                    key TEXT PRIMARY KEY,
                    value TEXT NOT NULL,
                    updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            ''')
            conn.commit()
        finally:
            conn.close()

def ensure_eventi_sistema_table():
    conn = get_db_connection()
    if conn:
        try:
            cursor = conn.cursor()
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS eventi_sistema (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    tipo_evento TEXT,
                    livello TEXT,
                    messaggio TEXT,
                    componente TEXT,
                    timestamp DATETIME DEFAULT CURRENT_TIMESTAMP
                )
            ''')
            conn.commit()
        finally:
            conn.close()

# Ora puoi chiamarla!
config_manager = get_config_manager()
ensure_system_settings_table()
ensure_eventi_sistema_table()

def check_main_process():
    """Controlla se il processo main.py è in esecuzione"""
    for proc in psutil.process_iter(['name', 'cmdline']):
        try:
            if proc.info['cmdline']:
                cmdline = ' '.join(proc.info['cmdline'])
                if 'python' in cmdline and 'main.py' in cmdline:
                    return True
        except (psutil.NoSuchProcess, psutil.AccessDenied):
            pass
    return False

def start_main_process():
    """Avvia il processo main.py in background"""
    try:
        main_path = os.path.join(project_root, 'src', 'main.py')
        subprocess.Popen(['python3', main_path], 
                        stdout=subprocess.PIPE,
                        stderr=subprocess.PIPE)
        logger.info("✅ Processo main.py avviato in background")
        return True
    except Exception as e:
        logger.error(f"❌ Errore avvio main.py: {e}")
        return False

# ===============================
# CARD READER E ODOO SYNC
# ===============================

# Variabili globali per il lettore di tessere e Odoo
card_reader = None
card_reader_thread = None
card_reader_running = False
odoo_connector = None
odoo_sync_thread = None
odoo_sync_running = False

def handle_card_read(codice_fiscale):
    """
    Gestisce la lettura di una tessera dal lettore Omnikey.
    Questa funzione viene chiamata dal CardReader quando viene letta una tessera.
    """
    logger.info(f"Tessera letta dal lettore Omnikey: {codice_fiscale[:4]}***{codice_fiscale[-4:]}")
    
    # Processa il codice fiscale
    result = process_codice_fiscale(codice_fiscale)
    
    # Gestisci l'hardware usando la configurazione dinamica del database
    try:
        controller = USBRLY08Controller()
        if controller.connect():
            try:
                # Recupera configurazione relè dal database
                conn = get_db_connection()
                if conn:
                    try:
                        cursor = conn.cursor()
                        cursor.execute("""
                            SELECT relay_number, description, valid_action, valid_duration, 
                                   invalid_action, invalid_duration 
                            FROM relay_config
                            ORDER BY relay_number
                        """)
                        configs = cursor.fetchall()
                        
                        # Applica configurazione in base al risultato
                        for relay_num, description, valid_action, valid_duration, invalid_action, invalid_duration in configs:
                            if result['authorized']:
                                # Applica azione per CF valido
                                if valid_action != 'OFF':
                                    logger.info(f"CF valido - Relè {relay_num} ({description}): {valid_action} per {valid_duration}s")
                                    
                                    if valid_action == 'ON':
                                        controller._send_command(100 + relay_num)  # Accendi
                                        if valid_duration > 0:
                                            threading.Timer(valid_duration, lambda rn=relay_num, ctrl=controller: ctrl._send_command(110 + rn)).start()
                                    elif valid_action == 'PULSE':
                                        controller._send_command(100 + relay_num)  # Accendi
                                        duration = valid_duration if valid_duration > 0 else 0.5
                                        threading.Timer(duration, lambda rn=relay_num, ctrl=controller: ctrl._send_command(110 + rn)).start()
                            else:
                                # Applica azione per CF non valido
                                if invalid_action != 'OFF':
                                    logger.info(f"CF non valido - Relè {relay_num} ({description}): {invalid_action} per {invalid_duration}s")
                                    
                                    if invalid_action == 'ON':
                                        controller._send_command(100 + relay_num)  # Accendi
                                        if invalid_duration > 0:
                                            threading.Timer(invalid_duration, lambda rn=relay_num, ctrl=controller: ctrl._send_command(110 + rn)).start()
                                    elif invalid_action == 'PULSE':
                                        controller._send_command(100 + relay_num)  # Accendi
                                        duration = invalid_duration if invalid_duration > 0 else 0.5
                                        threading.Timer(duration, lambda rn=relay_num, ctrl=controller: ctrl._send_command(110 + rn)).start()
                        
                        # Mantieni la connessione aperta per permettere ai timer di eseguire
                        # Si disconnetterà automaticamente dopo il tempo più lungo configurato
                        max_duration = max([cfg[3] if result['authorized'] else cfg[5] for cfg in configs], default=0)
                        if max_duration > 0:
                            threading.Timer(max_duration + 1.0, controller.disconnect).start()
                        else:
                            threading.Timer(5.0, controller.disconnect).start()
                        
                        if result['authorized']:
                            logger.info(f"Accesso autorizzato - Azioni relè eseguite")
                        else:
                            logger.warning(f"Accesso negato: {result['error_message']}")
                            
                    finally:
                        conn.close()
                else:
                    # Fallback alla configurazione di default se DB non disponibile
                    if result['authorized']:
                        controller.access_granted()
                        controller.open_gate(8.0)
                        logger.info(f"Cancello aperto per accesso autorizzato (config default)")
                    else:
                        controller.access_denied()
                        logger.warning(f"Accesso negato (config default): {result['error_message']}")
                    threading.Timer(10.0, controller.disconnect).start()
            except Exception as e:
                controller.disconnect()
                raise e
        else:
            logger.error("Errore connessione controller USB-RLY08")
    except Exception as e:
        logger.error(f"Errore hardware durante gestione tessera: {str(e)}")

def start_card_reader():
    """Avvia il lettore di tessere in un thread separato"""
    global card_reader, card_reader_thread, card_reader_running
    
    if card_reader_running:
        logger.warning("Lettore tessere già in esecuzione")
        return
    
    try:
        # Prima prova con CRT-285
        try:
            from hardware.crt285_reader import CRT285Reader
            logger.info("Avvio lettore tessere CRT-285...")
            # CRT-285 usa libusb direttamente, non una porta seriale
            card_reader = CRT285Reader(device_path=None, auto_test=False)
            # CRT285Reader si connette automaticamente nel costruttore
            if card_reader.lib:  # Verifica se la libreria è stata caricata
                logger.info("✅ CRT-285 connesso con successo")
                card_reader_running = True
            else:
                raise Exception("CRT-285 non disponibile")
        except Exception as e:
            # Fallback a Omnikey
            logger.warning(f"CRT-285 non disponibile: {e}, provo con Omnikey...")
            from hardware.card_reader import CardReader
            card_reader = CardReader()
            card_reader_running = True
        
        # Avvia il lettore in un thread separato
        card_reader_thread = threading.Thread(
            target=card_reader.start_continuous_reading,
            args=(handle_card_read,),
            daemon=True
        )
        card_reader_thread.start()
        
        logger.info("Lettore tessere Omnikey avviato con successo")
    except Exception as e:
        logger.error(f"Errore avvio lettore tessere: {str(e)}")
        card_reader_running = False

def stop_card_reader():
    """Ferma il lettore di tessere"""
    global card_reader, card_reader_running
    
    if not card_reader_running:
        return
    
    try:
        logger.info("Arresto lettore tessere...")
        if card_reader:
            card_reader.stop()
        card_reader_running = False
        logger.info("Lettore tessere arrestato")
    except Exception as e:
        logger.error(f"Errore arresto lettore tessere: {str(e)}")

# Funzioni per la sincronizzazione con Odoo
def configure_odoo_connector():
    """Configura il connettore Odoo"""
    global odoo_connector
    
    try:
        logger.info("Configurazione connettore Odoo...")
        
        # Crea un mock config manager per OdooPartnerConnector
        class MockConfigManager:
            def __init__(self):
                self.config = MockConfig()
            def get_config(self):
                return self.config
        
        class MockConfig:
            def __init__(self):
                self.odoo = None
        
        # Configurazione Odoo
        odoo_config = {
            'url': 'https://app.calabramaceri.it',
            'database': 'cmapp',
            'username': 'controllo-accessi@calabramaceri.it',
            'password': 'AcC3ss0C0ntr0l!2025#Rnd',
            'comune': 'Rende',
            'sync_interval_hours': 12
        }
        
        # Inizializza il connettore
        mock_config_manager = MockConfigManager()
        odoo_connector = OdooPartnerConnector(mock_config_manager)
        
        # Configura le credenziali
        odoo_connector.url = odoo_config['url']
        odoo_connector.db = odoo_config['database']
        odoo_connector.username = odoo_config['username']
        odoo_connector.password = odoo_config['password']
        
        # Cerca di autenticarsi
        if odoo_connector.authenticate():
            logger.info("✅ Connessione Odoo stabilita")
            # Avvia sync automatico in thread
            start_odoo_sync(odoo_config['sync_interval_hours'])
        else:
            logger.warning("⚠️ Impossibile connettersi a Odoo - sync disabilitato")
            
    except Exception as e:
        logger.error(f"Errore configurazione Odoo: {e}")
        logger.warning("Sync Odoo disabilitato")

def start_odoo_sync(interval_hours=12):
    """Avvia la sincronizzazione periodica con Odoo"""
    global odoo_sync_thread, odoo_sync_running
    
    if odoo_sync_running:
        logger.warning("Sync Odoo già in esecuzione")
        return
    
    def sync_loop():
        while odoo_sync_running:
            try:
                logger.info("🔄 Avvio sincronizzazione con Odoo...")
                success, stats = perform_odoo_sync()
                if success:
                    logger.info(f"✅ Sync completato: {stats}")
                else:
                    logger.error(f"❌ Sync fallito: {stats}")
            except Exception as e:
                logger.error(f"Errore durante sync: {e}")
            
            # Attendi il prossimo sync
            for i in range(interval_hours * 3600):
                if not odoo_sync_running:
                    break
                time.sleep(1)
    
    odoo_sync_running = True
    odoo_sync_thread = threading.Thread(target=sync_loop, daemon=True)
    odoo_sync_thread.start()
    logger.info(f"Sync Odoo avviato (ogni {interval_hours} ore)")

def stop_odoo_sync():
    """Ferma la sincronizzazione con Odoo"""
    global odoo_sync_running
    odoo_sync_running = False
    logger.info("Sync Odoo fermato")

def perform_odoo_sync():
    """Esegue la sincronizzazione con Odoo"""
    global odoo_connector
    
    if not odoo_connector:
        logger.warning("Connettore Odoo non configurato")
        return False, {"error": "Connettore Odoo non configurato"}
    
    try:
        logger.info("Esecuzione sincronizzazione Odoo...")
        
        # Ottieni connessione database
        conn = get_db_connection()
        if not conn:
            logger.error("Database non disponibile per sync Odoo")
            return False, {"error": "Database non disponibile"}
        
        # Crea un wrapper per DatabaseManager
        class DatabaseWrapper:
            def __init__(self, conn):
                self.conn = conn
            
            def verify_access(self, codice_fiscale):
                try:
                    cursor = self.conn.cursor()
                    cursor.execute('''
                        SELECT id, nome, note
                        FROM utenti_autorizzati 
                        WHERE codice_fiscale = ? AND attivo = 1
                    ''', (codice_fiscale.upper(),))
                    
                    result = cursor.fetchone()
                    
                    if result:
                        user_data = {
                            'id': result[0],
                            'nome': result[1],
                            'note': result[2]
                        }
                        return True, user_data
                    else:
                        return False, None
                except Exception as e:
                    logger.error(f"Errore verifica accesso: {e}")
                    return False, None
            
            def add_user(self, codice_fiscale, nome, note=None, creato_da=None, modificato_da=None):
                try:
                    cursor = self.conn.cursor()
                    cursor.execute('''
                        INSERT INTO utenti_autorizzati
                        (codice_fiscale, nome, note, attivo, creato_da, modificato_da)
                        VALUES (?, ?, ?, 1, ?, ?)
                    ''', (codice_fiscale.upper(), nome, note, creato_da or 'ODOO_SYNC', modificato_da or 'ODOO_SYNC'))
                    
                    self.conn.commit()
                    return True
                except Exception as e:
                    logger.error(f"Errore aggiunta utente: {e}")
                    return False
        
        db_wrapper = DatabaseWrapper(conn)
        
        # Esegui sync
        stats = odoo_connector.sync_citizens(db_wrapper, "Rende", test_mode=False)
        
        conn.close()
        
        logger.info(f"Sync completato: {stats}")
        return True, stats
        
    except Exception as e:
        logger.error(f"Errore sync Odoo: {e}")
        return False, {"error": str(e)}

# ===============================
# MAIN
# ===============================

if __name__ == '__main__':
    print("🌐 Avvio dashboard web modulare")
    print(f"📄 Database: {DB_PATH}")
    print("🔗 URL: http://0.0.0.0:5000")
    
    # Controlla e avvia main.py se necessario
    if not check_main_process():
        print("⚠️ Processo main.py non trovato, avvio in background...")
        if start_main_process():
            print("✅ Processo main.py avviato correttamente")
        else:
            print("❌ Errore avvio main.py")
    
    # Crea directory static se non esiste
    static_dir = os.path.join(os.path.dirname(__file__), 'static')
    if not os.path.exists(static_dir):
        os.makedirs(static_dir)
        print(f"📁 Creata directory static: {static_dir}")
    
    # Avvia il lettore di tessere PRIMA di avviare il server
    print("💳 Avvio lettore tessere...")
    start_card_reader()
    
    # Configura e avvia la sincronizzazione con Odoo
    print("🔄 Configurazione connettore Odoo...")
    configure_odoo_connector()
    
    # Registra funzioni di pulizia all'uscita
    import atexit
    atexit.register(stop_card_reader)
    atexit.register(stop_odoo_sync)
    
    # Avvia il server sulla porta 5000 fissa
    print(f"🔌 Porta selezionata: 5000")
    app.run(host='0.0.0.0', port=5000, debug=False, threaded=True)

print("=== ENDPOINTS REGISTRATI ===")
for rule in app.url_map.iter_rules():
    print(rule)
print("============================")

# Funzioni per la sincronizzazione con Odoo
def configure_odoo_connector():
    """Configura il connettore Odoo"""
    global odoo_connector
    
    try:
        logger.info("Configurazione connettore Odoo...")
        
        # Crea un mock config manager per OdooPartnerConnector
        class MockConfigManager:
            def __init__(self):
                self.config = MockConfig()
            def get_config(self):
                return self.config
        
        class MockConfig:
            def __init__(self):
                self.odoo = None
        
        # Configurazione Odoo
        odoo_config = {
            'url': 'https://app.calabramaceri.it',
            'database': 'cmapp',
            'username': 'controllo-accessi@calabramaceri.it',
            'password': 'AcC3ss0C0ntr0l!2025#Rnd',
            'comune': 'Rende',
            'sync_interval_hours': 12
        }
        
        # Inizializza il connettore
        mock_config_manager = MockConfigManager()
        odoo_connector = OdooPartnerConnector(mock_config_manager)
        odoo_connector.configure_connection(
            url=odoo_config['url'],
            database=odoo_config['database'],
            username=odoo_config['username'],
            password=odoo_config['password'],
            comune=odoo_config['comune'],
            sync_interval=odoo_config['sync_interval_hours'] * 3600
        )
        
        # Test connessione
        success, message = odoo_connector.test_connection()
        if success:
            logger.info(f"Connessione Odoo stabilita: {message}")
            
            # Avvia sync automatica
            start_odoo_sync()
            return True
        else:
            logger.warning(f"Test connessione Odoo fallito: {message}")
            return False
            
    except Exception as e:
        logger.error(f"Errore configurazione Odoo: {e}")
        return False

def start_odoo_sync():
    """Avvia la sincronizzazione automatica con Odoo"""
    global odoo_connector, odoo_sync_thread, odoo_sync_running
    
    if odoo_sync_running:
        logger.warning("Sincronizzazione Odoo già in esecuzione")
        return
    
    try:
        logger.info("Avvio sincronizzazione automatica Odoo...")
        odoo_sync_running = True
        
        # Esegui sync iniziale
        perform_odoo_sync()
        
        # Avvia thread per sync periodica
        def sync_worker():
            while odoo_sync_running:
                try:
                    # Attendi intervallo configurato (default 12 ore)
                    sleep_hours = 12
                    logger.info(f"Prossima sync Odoo in {sleep_hours} ore")
                    time.sleep(sleep_hours * 3600)
                    
                    if odoo_sync_running:
                        perform_odoo_sync()
                        
                except Exception as e:
                    logger.error(f"Errore sync worker Odoo: {e}")
                    time.sleep(3600)  # Riprova tra un'ora in caso di errore
        
        odoo_sync_thread = threading.Thread(target=sync_worker, daemon=True)
        odoo_sync_thread.start()
        
        logger.info("Sincronizzazione automatica Odoo avviata")
        
    except Exception as e:
        logger.error(f"Errore avvio sync Odoo: {e}")
        odoo_sync_running = False

def stop_odoo_sync():
    """Ferma la sincronizzazione automatica con Odoo"""
    global odoo_sync_running
    
    if not odoo_sync_running:
        return
    
    try:
        logger.info("Arresto sincronizzazione Odoo...")
        odoo_sync_running = False
        
        if odoo_connector:
            odoo_connector.disconnect()
            
        logger.info("Sincronizzazione Odoo arrestata")
        
    except Exception as e:
        logger.error(f"Errore arresto sync Odoo: {e}")

def perform_odoo_sync():
    """Esegue la sincronizzazione con Odoo"""
    global odoo_connector
    
    if not odoo_connector:
        logger.warning("Connettore Odoo non configurato")
        return False, {"error": "Connettore Odoo non configurato"}
    
    try:
        logger.info("Esecuzione sincronizzazione Odoo...")
        
        # Ottieni connessione database
        conn = get_db_connection()
        if not conn:
            logger.error("Database non disponibile per sync Odoo")
            return False, {"error": "Database non disponibile"}
        
        # Crea un wrapper per DatabaseManager
        class DatabaseWrapper:
            def __init__(self, conn):
                self.conn = conn
            
            def verify_access(self, codice_fiscale):
                try:
                    cursor = self.conn.cursor()
                    cursor.execute('''
                        SELECT id, nome, note
                        FROM utenti_autorizzati 
                        WHERE codice_fiscale = ? AND attivo = 1
                    ''', (codice_fiscale.upper(),))
                    
                    result = cursor.fetchone()
                    
                    if result:
                        user_data = {
                            'id': result[0],
                            'nome': result[1],
                            'note': result[2]
                        }
                        return True, user_data
                    else:
                        return False, None
                except Exception as e:
                    logger.error(f"Errore verifica accesso: {e}")
                    return False, None
            
            def add_user(self, codice_fiscale, nome, note=None, creato_da=None, modificato_da=None):
                try:
                    cursor = self.conn.cursor()
                    cursor.execute('''
                        INSERT INTO utenti_autorizzati 
                        (codice_fiscale, nome, note, creato_da, modificato_da)
                        VALUES (?, ?, ?, ?, ?)
                    ''', (
                        codice_fiscale.upper(),
                        nome,
                        note,
                        creato_da,
                        modificato_da
                    ))
                    self.conn.commit()
                    
                    logger.info(f"Utente aggiunto: {nome} ({codice_fiscale})")
                    return True
                    
                except sqlite3.IntegrityError:
                    logger.warning(f"Utente già esistente: {codice_fiscale}")
                    return False
                except Exception as e:
                    logger.error(f"Errore aggiunta utente: {e}")
                    return False
        
        # Esegui sincronizzazione
        db_wrapper = DatabaseWrapper(conn)
        success, stats = odoo_connector.sync_to_database(db_wrapper)
        
        if success:
            logger.info(f"Sync Odoo completata: {stats['added']} cittadini aggiunti")
            return True, stats
        else:
            logger.warning("Sync Odoo fallita")
            return False, stats
            
    except Exception as e:
        logger.error(f"Errore sync Odoo: {e}")
        return False, {"error": str(e)}

# Registra funzione di pulizia all'uscita
import atexit
try:
    atexit.register(stop_card_reader)
except NameError:
    logger.warning("stop_card_reader non definita, skip registrazione atexit")
try:
    atexit.register(stop_odoo_sync)
except NameError:
    logger.warning("stop_odoo_sync non definita, skip registrazione atexit")

def scheduled_backup():
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    backup_filename = f'access_{timestamp}.db'
    src = DB_PATH
    dst = os.path.join(BACKUP_DIR, backup_filename)
    shutil.copy2(src, dst)
    print(f"[SCHEDULER] Backup creato: {backup_filename}")

    # Retention
    conn = get_db_connection()
    retention = 7
    if conn:
        try:
            cursor = conn.cursor()
            cursor.execute("SELECT value FROM system_settings WHERE key='backup.retention'")
            row = cursor.fetchone()
            if row:
                retention = int(row[0])
        finally:
            conn.close()
    files = sorted(
        [os.path.join(BACKUP_DIR, f) for f in os.listdir(BACKUP_DIR) if os.path.isfile(os.path.join(BACKUP_DIR, f))],
        key=lambda x: os.path.getmtime(x)
    )
    to_delete = files[:-retention]
    for f in to_delete:
        os.remove(f)
        print(f"[SCHEDULER] Backup eliminato per retention: {os.path.basename(f)}")

# Scarica backup
@app.route('/api/backup/download/<filename>')
@require_auth()
@require_permission('all')
def api_backup_download(filename: str) -> FlaskResponse:
    return backup_module.download_backup(filename)

# Elimina backup
@app.route('/api/backup/delete/<filename>', methods=['DELETE'])
@require_auth()
@require_permission('all')
def api_backup_delete(filename: str) -> FlaskResponse:
    return backup_module.delete_backup(filename)

# Ripristina backup (sovrascrive access.db)
@app.route('/api/backup/restore/<filename>', methods=['POST'])
@require_auth()
@require_permission('all')
def api_backup_restore(filename: str) -> FlaskResponse:
    return backup_module.restore_backup(filename)
